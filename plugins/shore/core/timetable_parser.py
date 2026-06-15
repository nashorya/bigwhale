"""
课表解析工具 — 支持 xlsx 和图片课表。

从 generate_schedule_csv.py 移植并适配 Bot 异步架构。
解析课表后提取空闲时段描述，供 AI 排课 prompt 使用。
"""

from __future__ import annotations

import json
import logging
from typing import Any

logger = logging.getLogger("shore.timetable")

# 被视为"空闲"的课表单元格内容
_FREE_CELL_VALUES = {"", "自习", "None", "-", "——", "none", "null"}


# ──────────────────────────────────────────────
# xlsx 课表解析
# ──────────────────────────────────────────────


def parse_xlsx_timetable(xlsx_path: str) -> dict[str, Any]:
    """
    解析 xlsx 格式的课表文件，返回结构化课表数据。

    返回格式：
        {
            "headers": ["", "周一", "周二", ...],
            "slots": ["第1-2节", "第3-4节", ...],
            "busy": [{"day": "周一", "slot": "第1-2节", "course": "高数"}]
        }
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("需要安装 openpyxl：pip install openpyxl")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    # 处理合并单元格
    merge_values: dict[tuple[int, int], Any] = {}
    for rng in ws.merged_cells.ranges:
        val = ws.cell(rng.min_row, rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                merge_values[(r, c)] = val

    def get_cell(r: int, c: int) -> str:
        v = merge_values.get((r, c), ws.cell(r, c).value)
        return str(v).strip() if v is not None else ""

    # 读取所有行（跳过空行）
    rows = [
        row
        for row in [
            [get_cell(r, c) for c in range(1, ws.max_column + 1)]
            for r in range(1, ws.max_row + 1)
        ]
        if any(row)
    ]

    if not rows:
        return {"headers": [], "slots": [], "busy": []}

    headers = rows[0]
    busy: list[dict[str, str]] = []

    for r in rows[1:]:
        slot = r[0] if r else ""
        for i, cell in enumerate(r[1:], 1):
            if cell and cell not in _FREE_CELL_VALUES:
                day = headers[i] if i < len(headers) else f"列{i}"
                busy.append({"day": day, "slot": slot, "course": cell})

    return {
        "headers": headers,
        "slots": [r[0] for r in rows[1:] if r],
        "busy": busy,
    }


# ──────────────────────────────────────────────
# 图片课表解析（多模态模型）
# ──────────────────────────────────────────────


async def parse_image_timetable(
    image_bytes: bytes, mime_type: str = "image/jpeg"
) -> dict[str, Any]:
    """
    使用 OpenAI 兼容协议的多模态模型识别图片课表，返回结构化数据。

    参数：
        image_bytes: 图片二进制数据
        mime_type: 图片 MIME 类型

    返回：
        同 parse_xlsx_timetable 的格式
    """
    import base64
    import os

    from openai import AsyncOpenAI

    api_key = os.environ.get("API_KEY", "")
    if not api_key:
        raise RuntimeError("环境变量 API_KEY 未设置，请在 .env 中配置 API Key。")
    model = os.environ.get("CHAT_MODEL", "")
    if not model:
        raise RuntimeError("环境变量 CHAT_MODEL 未设置，请在 .env 中配置模型名。")

    client_kwargs: dict[str, Any] = {"api_key": api_key}
    base_url = os.environ.get("BASE_URL", "")
    if base_url:
        client_kwargs["base_url"] = base_url
    client = AsyncOpenAI(**client_kwargs)

    # 构建多模态请求
    prompt = (
        "这是一张大学课程表。请识别其中每天每节课的上课安排。\n"
        "返回 JSON 对象，格式如下：\n"
        "{\n"
        '  "headers": ["", "周一", "周二", "周三", "周四", "周五"],\n'
        '  "slots": ["第1-2节(08:00-09:35)", ...],\n'
        '  "busy": [\n'
        '    {"day": "周一", "slot": "第1-2节", "course": "高等数学"},\n'
        "    ...\n"
        "  ]\n"
        "}\n"
        "只返回 JSON 对象，不要其他文字。空白和自习格不要放入 busy。"
    )

    image_data = base64.b64encode(image_bytes).decode("ascii")
    response = await client.chat.completions.create(
        model=model,
        messages=[
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:{mime_type};base64,{image_data}",
                        },
                    },
                ],
            }
        ],
        temperature=0.1,
        max_tokens=4000,
    )

    text = response.choices[0].message.content or ""

    # 清理 JSON
    import re

    md_match = re.search(r"```(?:json)?\s*\n?(.*?)```", text, re.DOTALL)
    if md_match:
        text = md_match.group(1).strip()

    result = json.loads(text)
    logger.info("图片课表识别成功：%d 个有课时段", len(result.get("busy", [])))
    return result


# ──────────────────────────────────────────────
# 空闲时段提取
# ──────────────────────────────────────────────


def format_busy_desc(timetable: dict[str, Any]) -> str:
    """
    将结构化课表转为人类可读的「有课时段」描述，
    供 LLM prompt 注入。

    返回示例：
        周一 第1-2节：高等数学
        周一 第5-6节：英语
        周三 第3-4节：线性代数
        ...
    """
    busy = timetable.get("busy", [])
    if not busy:
        return ""

    lines = []
    for b in busy:
        lines.append(f"  {b['day']} {b['slot']}：{b['course']}")

    return "\n".join(lines)


def make_timetable_prompt_section(timetable: dict[str, Any], months: str = "") -> str:
    """
    生成课表约束的 prompt 段落，直接插入排课 prompt。

    参数：
        timetable: 结构化课表数据
        months: 课表适用月份（如 "3-6月"）

    返回：
        可直接插入 prompt 的课表段落文本
    """
    busy_desc = format_busy_desc(timetable)
    if not busy_desc:
        return ""

    all_slots = "、".join(timetable.get("slots", []))
    month_note = f"（适用月份：{months}）" if months else ""

    return (
        f"\n## 大学课表约束{month_note}\n"
        f"以下时段有大学课程，绝对不能安排考研复习：\n"
        f"{busy_desc}\n"
        f"\n所有节次：{all_slots}\n"
        f"请只在空闲节次安排考研学习。周末默认无课，全天可安排。\n"
    )
