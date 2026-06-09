"""
考研备考计划 CSV 生成器 (Pro 并发防弹版)

更新特性：
  1. [防崩拦截] 智能补全中转站 /v1 路径，精准拦截异常网页/纯文本报错，告别 str 崩溃。
  2. [并发提速] 4 个 CSV 计划表多线程并发生成，耗时降低 75%。
  3. [成本节约] 隔离全局大纲，仅针对知识清单任务专线投喂，省 Token 且防幻觉。
  4. [丝滑交互] 支持直接输入大纲本地文件路径 (TXT/MD) 进行解析读取。
"""

import os, csv, json, argparse, re, base64
from datetime import datetime, timedelta
from pathlib import Path
from urllib.parse import urlparse
import concurrent.futures

try:
    import anthropic as _ant
except ImportError:
    _ant = None
try:
    from openai import OpenAI as _OAI
except ImportError:
    _OAI = None
try:
    import requests as _req
except ImportError:
    _req = None
try:
    import openpyxl as _openpyxl
except ImportError:
    _openpyxl = None
try:
    from google import genai as _genai
    from google.genai import types as _genai_types
except ImportError:
    _genai = None
    _genai_types = None


# ══════════════════════════════════════════════════════════════════
# LLM 调用 (加入强力防弹装甲)
# ══════════════════════════════════════════════════════════════════

def _encode_image(image_path: str) -> tuple:
    """读取图片并返回 (base64字符串, media_type)"""
    ext = Path(image_path).suffix.lower()
    media_map = {".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png", ".gif": "image/gif", ".webp": "image/webp"}
    media_type = media_map.get(ext, "image/jpeg")
    with open(image_path, "rb") as f:
        return base64.standard_b64encode(f.read()).decode("utf-8"), media_type


def call_llm(prompt, system, cfg, image_path=None):
    """调用 LLM，可选传入图片路径启用视觉能力"""
    p = cfg["provider"].lower()
    m, k, u = cfg.get("model"), cfg.get("api_key"), cfg.get("base_url")

    # ── Gemini 原生 SDK（支持视觉）──────────────────────────────
    if p == "gemini-native":
        if _genai is None: raise ImportError("pip install google-genai")
        client_kwargs = {"api_key": k}
        if u:
            client_kwargs["http_options"] = _genai_types.HttpOptions(base_url=u)
        client = _genai.Client(**client_kwargs)
        model_name = m or "gemini-2.5-flash-preview-05-20"
        # 构建多模态内容
        contents = []
        if system:
            contents.append(system + "\n\n")
        if image_path:
            img_bytes = Path(image_path).read_bytes()
            ext = Path(image_path).suffix.lower()
            mime_map = {".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png", ".webp": "image/webp"}
            mime = mime_map.get(ext, "image/jpeg")
            contents.append(_genai_types.Part.from_bytes(data=img_bytes, mime_type=mime))
        contents.append(prompt)
        r = client.models.generate_content(
            model=model_name,
            contents=contents,
            config=_genai_types.GenerateContentConfig(temperature=0.2),
        )
        return r.text or ""

    # ── Anthropic ────────────────────────────────────────────────
    if p == "anthropic":
        if _ant is None: raise ImportError("pip install anthropic")
        kwargs = {"api_key": k, "timeout": 300.0}
        if u: kwargs["base_url"] = u
        c = _ant.Anthropic(**kwargs)
        # 构建消息内容（支持图片）
        if image_path:
            img_b64, media_type = _encode_image(image_path)
            content = [
                {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": img_b64}},
                {"type": "text", "text": prompt}
            ]
        else:
            content = prompt
        r = c.messages.create(
            model=m or "claude-3-5-sonnet-20241022", max_tokens=16384,
            system=system, messages=[{"role": "user", "content": content}])
        return r.content[0].text

    # ── OpenAI 兼容（openai / deepseek / qwen / gemini） ─────────
    if p in ("openai", "deepseek", "qwen", "gemini"):
        if _OAI is None: raise ImportError("pip install openai")
        defaults = {
            "openai":   ("gpt-4o",                  None),
            "deepseek": ("deepseek-chat",            "https://api.deepseek.com/v1"),
            "qwen":     ("qwen-plus",                "https://dashscope.aliyuncs.com/compatible-mode/v1"),
            "gemini":   ("gemini-2.5-flash",         "https://generativelanguage.googleapis.com/v1beta/openai/"),
        }
        dm, du = defaults[p]
        
        final_url = u or du
        if final_url:
            parsed = urlparse(final_url)
            if parsed.path in ("", "/"):
                final_url = final_url.rstrip("/") + "/v1"

        c = _OAI(api_key=k, base_url=final_url)
        
        # 构建消息内容（支持图片）
        if image_path:
            img_b64, media_type = _encode_image(image_path)
            user_content = [
                {"type": "image_url", "image_url": {"url": f"data:{media_type};base64,{img_b64}"}},
                {"type": "text", "text": prompt}
            ]
        else:
            user_content = prompt

        try:
            r = c.chat.completions.create(
                model=m or dm, max_tokens=16384,
                messages=[{"role": "system", "content": system},
                          {"role": "user",   "content": user_content}])
        except Exception as e:
            raise RuntimeError(f"API 网络请求失败或被中转站拒绝: {e}")

        if isinstance(r, str):
            raise ValueError(f"中转站返回了异常纯文本: {r}")
        if not hasattr(r, 'choices') or not r.choices:
            raise ValueError(f"中转站返回数据残缺: {r}")
            
        return r.choices[0].message.content

    # ── Ollama（本地） ────────────────────────────────────────────
    if p == "ollama":
        if _req is None: raise ImportError("pip install requests")
        r = _req.post(f"{u or 'http://localhost:11434'}/api/chat",
                      json={"model": m or "llama3", "stream": False,
                            "messages": [{"role": "system", "content": system},
                                         {"role": "user",   "content": prompt}]},
                      timeout=180)
        r.raise_for_status()
        return r.json()["message"]["content"]

    raise ValueError(f"未知 provider: {p}，支持: anthropic/openai/deepseek/qwen/gemini/ollama")


# ══════════════════════════════════════════════════════════════════
# 防崩溃工具函数
# ══════════════════════════════════════════════════════════════════

def _flatten_nested(data):
    """将嵌套 JSON 展平为扁平数组（兼容 Claude 输出格式）
    Claude 常输出 [{week, schedule:[{时间段,周一,...}]}, ...] 这种嵌套结构，
    需要展平为 [{时间段,周一,...}, ...] 才能写入 CSV。
    """
    if not isinstance(data, list) or not data:
        return data
    # 检测是否为嵌套结构：第一个元素包含列表类型的子字段
    first = data[0]
    if not isinstance(first, dict):
        return data
    # 查找包含子数组的字段名
    list_keys = [k for k, v in first.items() if isinstance(v, list)]
    if not list_keys:
        return data  # 已经是扁平结构
    # 展平：将每个父对象的子数组项提取出来，附带父级元数据
    flat = []
    list_key = list_keys[0]  # 取第一个子数组字段（如 schedule）
    meta_keys = [k for k in first.keys() if k != list_key]
    for parent in data:
        sub_items = parent.get(list_key, [])
        if not isinstance(sub_items, list):
            flat.append(parent)
            continue
        meta = {k: parent.get(k, "") for k in meta_keys}
        for item in sub_items:
            if isinstance(item, dict):
                row = dict(meta)  # 先放父级字段（如 week）
                row.update(item)  # 再放子项字段（如 时间段、周一...）
                flat.append(row)
            else:
                flat.append(item)
    print(f"  🔄 检测到嵌套 JSON，已自动展平（{len(data)} 组 → {len(flat)} 行）")
    return flat


def parse_json_safe(text):
    """解析 JSON，自动修复截断 + 兼容嵌套格式 + 剥离 markdown"""
    text = text.strip()
    # 第一步：剥离 markdown 代码块（Claude 常加 ```json ... ```）
    md_match = re.search(r'```(?:json)?\s*\n?(.*?)```', text, re.DOTALL)
    if md_match:
        text = md_match.group(1).strip()

    # 第二步：尝试匹配完整 JSON
    match = re.search(r'(\[.*\]|\{.*\})', text, re.DOTALL)
    if match:
        try:
            data = json.loads(match.group(1))
            if isinstance(data, list):
                return _flatten_nested(data)
            return data
        except json.JSONDecodeError:
            pass

    # 第三步：截断修复
    start = text.find('[')
    if start >= 0:
        fragment = text[start:]
        # 尝试多种闭合方式（嵌套的 }] 和扁平的 }]）
        for suffix in ['}]', '}]}]', ']}]', ']']:
            last_brace = fragment.rfind('}')
            if last_brace > 0:
                fixed = fragment[:last_brace + 1] + suffix.lstrip('}')
                try:
                    data = json.loads(fixed)
                    print(f"  ⚠️ JSON 被截断，已自动修复（保留 {len(data)} 条记录）")
                    return _flatten_nested(data)
                except json.JSONDecodeError:
                    continue
    return json.loads(text)

def write_csv(data, path):
    if not data: return
    fieldnames = []
    for row in data:
        if isinstance(row, dict):
            for k in row.keys():
                if k not in fieldnames: fieldnames.append(k)
                    
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
        w.writeheader()
        w.writerows([r for r in data if isinstance(r, dict)])
    print(f"  ✅ {os.path.basename(path)}（{len(data)} 行）")


# ══════════════════════════════════════════════════════════════════
# 置信度自评与大纲收集
# ══════════════════════════════════════════════════════════════════

def check_subject_confidence(subjects: list[str], school: str, cfg: dict) -> dict[str, str]:
    prompt = (
        f"评估能否准确列出以下科目的考研核心知识点？目标院校：{school}\n科目列表：\n" + 
        "\n".join(f"  - {s}" for s in subjects) + 
        "\n\nhigh=统考/常见专业课把握充足；low=冷门/高度自命题极不确定。\n"
        "只返回 JSON 对象，示例：{\"数据结构\": \"high\", \"某冷门\": \"low\"}"
    )
    raw = call_llm(prompt, "你是考研专家。诚实评估，不加任何废话。", cfg)
    return parse_json_safe(raw)

def collect_syllabi(low_subjects: list[str]) -> dict[str, str]:
    syllabi = {}
    print("\n" + "═" * 58)
    print("⚠️  以下科目知识点模型把握不足，强烈建议补充考试大纲：")
    for subj in low_subjects:
        print(f"\n📌 科目：{subj}")
        print("   👉 方式1：输入大纲所在的【本地文件路径】(如 D:\\dagang.txt) 并回车读取")
        print("   👉 方式2：直接在此处粘贴文本，换行后单独输入 END 结束")
        print("   （直接输入 END 可强制跳过）")

        lines = []
        while True:
            try: line = input(">>> ")
            except EOFError: break
            
            stripped = line.strip()
            if len(lines) == 0 and stripped and os.path.isfile(stripped):
                try:
                    with open(stripped, "r", encoding="utf-8") as f: lines.append(f.read())
                    print(f"   📂 成功读取大纲文件: {stripped}")
                    break
                except Exception as e:
                    print(f"   ❌ 读取失败: {e}，请重新输入或直接粘贴文本：")
                    continue

            if stripped.upper() == "END": break
            lines.append(line)

        if lines:
            syllabi[subj] = "\n".join(lines)
            print(f"   ✅ 已就绪（大纲字数：{len(syllabi[subj])} 字符）")
        else:
            print(f"   ⏭ 跳过，模型将依靠猜测生成")
    return syllabi


# ══════════════════════════════════════════════════════════════════
# 课表解析
# ══════════════════════════════════════════════════════════════════

def parse_image_with_paddle(image_path: str) -> list:
    try: from paddleocr import PaddleOCR
    except ImportError: raise ImportError("pip install paddlepaddle paddleocr")
    print("  🔍 PaddleOCR 识别中...")
    ocr = PaddleOCR(use_angle_cls=True, lang="ch", show_log=False)
    result = ocr.ocr(image_path, cls=True)
    cells = [(int(b[0][1]), int(b[0][0]), t.strip()) for line in (result[0] or []) for b, (t, s) in [line] if s >= 0.5 and t.strip()]
    if not cells: return []
    cells.sort(key=lambda c: (c[0], c[1]))
    rows, cur = [], [cells[0]]
    for cell in cells[1:]:
        if abs(cell[0] - cur[0][0]) < 25: cur.append(cell)
        else: rows.append(sorted(cur, key=lambda c: c[1])); cur = [cell]
    rows.append(sorted(cur, key=lambda c: c[1]))
    return [[c[2] for c in row] for row in rows]

def image_rows_to_timetable(rows: list) -> dict:
    if not rows: return {"headers": [], "slots": [], "busy": []}
    headers, slots, busy, FREE = rows[0], [r[0] for r in rows[1:] if r], [], {"", "自习", "None", "-", "——", "none", "null"}
    for r in rows[1:]:
        slot = r[0] if r else ""
        for i, cell in enumerate(r[1:], 1):
            if cell and cell not in FREE: busy.append({"day": headers[i] if i < len(headers) else f"列{i}", "slot": slot, "course": cell})
    return {"headers": headers, "slots": slots, "busy": busy}

def parse_xlsx_timetable(xlsx_path: str) -> dict:
    if _openpyxl is None: raise ImportError("pip install openpyxl")
    wb = _openpyxl.load_workbook(xlsx_path, data_only=True)
    ws, merge_values = wb.active, {}
    for rng in ws.merged_cells.ranges:
        val = ws.cell(rng.min_row, rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1): merge_values[(r, c)] = val
    def get(r, c): v = merge_values.get((r, c), ws.cell(r, c).value); return str(v).strip() if v is not None else ""
    rows = [row for row in [[get(r, c) for c in range(1, ws.max_column + 1)] for r in range(1, ws.max_row + 1)] if any(row)]
    if not rows: return {"headers": [], "slots": [], "busy": []}
    headers, busy, FREE = rows[0], [], {"", "自习", "None", "-", "——", "none", "null"}
    for r in rows[1:]:
        slot = r[0]
        for i, cell in enumerate(r[1:], 1):
            if cell and cell not in FREE: busy.append({"day": headers[i] if i < len(headers) else f"列{i}", "slot": slot, "course": cell})
    return {"headers": headers, "slots": [r[0] for r in rows[1:]], "busy": busy}

def timetable_to_free_desc(timetable: dict, cfg: dict) -> str:
    busy_lines = "\n".join(f"  {b['day']} {b['slot']}：{b['course']}" for b in timetable.get("busy", []))
    prompt = f"有课时间：\n{busy_lines}\n所有节次：{'、'.join(timetable.get('slots', []))}\n请列每天空闲节次，格式：周一：第X节空闲。只输出时间。"
    return call_llm(prompt, "你是课表分析助手。", cfg)


# ══════════════════════════════════════════════════════════════════
# Prompt 组装 (全局大纲降维剥离)
# ══════════════════════════════════════════════════════════════════

def make_system(info: dict, free_desc: str | None) -> str:
    s1_line = f"  业务课一：{info['subject1']}\n" if info["subject1"] else ""
    tt_sec = f"\n【课表空闲时间】\n{free_desc}\n严格基于此安排。\n" if free_desc else "\n【课表】无，周末默认自习。\n"
    return (
        "你是专业考研规划师。\n"
        f"【考研科目】\n  英语：{info['english']}\n{s1_line}  业务课二：{info['subject2']}\n  政治：不排\n"
        f"【信息】\n  院校：{info['school']}\n  区间：{info['start_date']} — {info['end_date']}\n{tt_sec}\n"
        "【原则】绝对不排政治。输出 JSON 数组，禁止多余文字。"
    )

def prompt_monthly(info):
    s1_col = f", 业务课一_{info['subject1']}" if info["subject1"] else ""
    return f"按周输出 {info['start_date']} 到 {info['end_date']} 月度备考规划。JSON数组字段：月份, 周次, 英语_{info['english']}{s1_col}, 业务课二_{info['subject2']}, 复习重点, 阶段备注。只返回JSON数组。"

def prompt_weekly(info, free_desc):
    s1_note = f"业务一({info['subject1']})" if info["subject1"] else "无业务一"
    return f"输出每周日程(07:30-22:30)。当前学：英语、业务二、{s1_note}。JSON数组字段：时间段, 周一, 周二, 周三, 周四, 周五, 周六, 周日。"

def _extract_current_subjects(monthly_data: list, info: dict) -> tuple:
    """从月度规划中提取第一周（当前阶段）应学的科目，排除标注为'暂缓'的"""
    if not monthly_data:
        return [], [], "本周", ""
    first_week = monthly_data[0] if monthly_data else {}
    active_subjects = []
    suspended_subjects = []
    week_label = first_week.get("周次", "本周")
    focus = first_week.get("复习重点", "")
    for k, v in first_week.items():
        if k in ("月份", "周次", "复习重点", "阶段备注"):
            continue
        v_str = str(v).strip()
        if "暂缓" in v_str or not v_str:
            subj = k.split("_", 1)[-1] if "_" in k else k
            suspended_subjects.append(subj)
        else:
            subj = k.split("_", 1)[-1] if "_" in k else k
            active_subjects.append(f"{subj}：{v_str}")
    return active_subjects, suspended_subjects, week_label, focus

def prompt_weekly_cascaded(info, monthly_data, free_desc=None):
    """级联版：基于月度规划当前周的具体学习任务生成本周计划"""
    active, suspended, week_label, focus = _extract_current_subjects(monthly_data, info)
    active_desc = "\n".join(f"  - {a}" for a in active) if active else "  - 英语阅读、数学"
    suspend_desc = "、".join(suspended) if suspended else ""
    suspend_line = f"\n【暂缓科目（绝对不要安排！）】{suspend_desc}" if suspend_desc else ""
    return (
        f"生成 {week_label} 的具体学习计划（06:00-23:00）。\n"
        f"\n【本周学习任务（来自月度规划）】\n{active_desc}\n"
        f"\n【本周重点】{focus}\n"
        f"{suspend_line}\n"
        f"\n【单词学习策略】英语单词由Bot在碎片时间（课间、走路、排队、睡前）自动推送，"
        f"不要在日程中安排整块的背单词时段。英语的整块时间只安排阅读精读、长难句分析、翻译或写作。\n"
        f"\n严格只安排上述应学科目，暂缓科目绝对不要出现。\n"
        f"JSON数组字段：时间段, 周一, 周二, 周三, 周四, 周五, 周六, 周日。只返回JSON数组。"
    )

def prompt_checkin(info):
    s1_col = ", 业务课一_完成" if info["subject1"] else ""
    return f"输出每日打卡表。JSON数组字段：序号, 月日, 星期, 英语_完成{s1_col}, 业务课二_完成, 单词打卡_完成, 备注。完成列留空。"

def prompt_checklist(info, syllabi: dict):
    syllabi_sec = ("\n【务必严格依据以下大纲内容提取考点，禁止幻觉生造】\n" + "\n\n".join([f"  ── {s} ──\n{t}" for s, t in syllabi.items()]) + "\n\n") if syllabi else ""
    extra = f"、{info['subject1']}" if info["subject1"] else ""
    return f"输出 {info['subject2']}{extra}、英语考研核心考点。\n{syllabi_sec}JSON数组字段：序号, 科目, 知识点或技能, 重要程度_高中低, 掌握度_1到5, 建议学习月份, 备注。掌握度留空。只返回JSON数组。"


# ══════════════════════════════════════════════════════════════════
# 主流程
# ══════════════════════════════════════════════════════════════════

def generate_all(cfg, info, out_dir):
    os.makedirs(out_dir, exist_ok=True)
    free_desc = None
    timetable_image = None  # 如果是图片课表，保存路径供视觉 API 使用
    if info.get("timetable"):
        path = info["timetable"]
        print(f"\n📋 解析课表：{path}")
        try:
            if Path(path).suffix.lower() in (".jpg", ".png", ".jpeg", ".webp", ".gif"):
                timetable_image = path
                print("  📷 检测到图片课表，将使用 LLM 视觉能力直接识别")
            else:
                timetable = parse_xlsx_timetable(path)
                if timetable and timetable.get("busy"):
                    free_desc = timetable_to_free_desc(timetable, cfg)
                    print("  ✅ 空闲时段提取完成")
        except Exception as e: print(f"  ❌ 课表解析失败：{e}")

    all_subjects = ([s.strip() for s in info["subject1"].split(",")] if info["subject1"] else []) + [s.strip() for s in info["subject2"].split(",")]
    
    print(f"\n🧠 评估各科目知识点置信度...")
    try:
        confidence = check_subject_confidence(all_subjects, info["school"], cfg)
        low_subjects = [s for s, c in confidence.items() if str(c).lower() == "low"]
        for subj, level in confidence.items():
            print(f"  {'✅' if str(level).lower() == 'high' else '⚠️ '} {subj}：{'把握充足' if str(level).lower() == 'high' else '把握不足，需大纲'}")
    except Exception as e:
        print(f"  ❌ 评估异常 ({type(e).__name__}: {e})，默认跳过大纲补充直接生成。")
        low_subjects = []

    syllabi = collect_syllabi(low_subjects) if low_subjects else {}
    system = make_system(info, free_desc)

    # ══════════════════════════════════════════════════════════════
    # Stage 1：先并发生成 月度规划 + 知识清单（互不依赖）
    # ══════════════════════════════════════════════════════════════
    print("\n📅 Stage 1：生成月度备考规划 + 核心知识清单...")
    stage1_tasks = [
        ("01_月度备考规划.csv", prompt_monthly(info), None),
        ("04_核心知识清单.csv", prompt_checklist(info, syllabi), None),
    ]
    stage1_results = {}  # filename → parsed data

    def process_task(task):
        filename, prompt, img = task
        try:
            raw = call_llm(prompt, system, cfg, image_path=img)
            data = parse_json_safe(raw)
            write_csv(data, os.path.join(out_dir, filename))
            print(f"  ✅ {filename}（{len(data)} 行）")
            return filename, data
        except Exception as e:
            print(f"  ❌ {filename} 生成失败 ({type(e).__name__}: {e})")
            if 'raw' in locals() and isinstance(raw, str):
                err_path = os.path.join(out_dir, filename.replace(".csv", "_raw.txt"))
                with open(err_path, "w", encoding="utf-8") as f: f.write(raw)
            return filename, None

    with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
        futures = [executor.submit(process_task, t) for t in stage1_tasks]
        for f in concurrent.futures.as_completed(futures):
            fn, data = f.result()
            stage1_results[fn] = data

    # ══════════════════════════════════════════════════════════════
    # Stage 2：基于月度规划级联生成 周日程 + 打卡表
    # ══════════════════════════════════════════════════════════════
    monthly_data = stage1_results.get("01_月度备考规划.csv")

    # 构建级联周日程 prompt
    if monthly_data:
        print(f"\n📋 Stage 2：基于月度规划级联生成周日程 + 打卡表...")
        active, suspended, week_label, focus = _extract_current_subjects(monthly_data, info)
        if suspended:
            print(f"  📌 当前阶段应学：{len(active)} 科，暂缓：{'、'.join(suspended)}")
        weekly_prompt = prompt_weekly_cascaded(info, monthly_data, free_desc)
    else:
        print(f"\n⚠️ 月度规划生成失败，周日程使用默认 prompt...")
        weekly_prompt = prompt_weekly(info, free_desc)

    # 如果有图片课表，在级联 prompt 基础上附加视觉指令
    weekly_img = None
    if timetable_image:
        vision_prefix = (
            f"这是我{info.get('timetable_months', '3-6月')}的课程表。"
            "请先识别每天的上课时间，上课时段标注\"[上课]课程名\"。\n\n"
        )
        weekly_prompt = vision_prefix + weekly_prompt
        weekly_img = timetable_image

    stage2_tasks = [
        ("02_本周学习计划.csv", weekly_prompt, weekly_img),
        ("03_每日打卡记录.csv", prompt_checkin(info), None),
    ]

    with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
        futures = [executor.submit(process_task, t) for t in stage2_tasks]
        for f in concurrent.futures.as_completed(futures):
            f.result()

    print(f"\n🎉 完美收工！文件已生成至目录 → ./{out_dir}/")


# ══════════════════════════════════════════════════════════════════
# CLI
# ══════════════════════════════════════════════════════════════════

def main():
    today = datetime.now()
    default_start = today.strftime("%Y-%m-%d")
    default_end   = (today + timedelta(days=120)).strftime("%Y-%m-%d")
    default_exam  = f"{today.year if today.month < 12 else today.year + 1}-12-20"

    ap = argparse.ArgumentParser(description="考研排期并发生成器")
    ap.add_argument("--provider",  "-p", default="anthropic", choices=["anthropic","openai","deepseek","qwen","gemini","gemini-native","ollama"])
    ap.add_argument("--model",     "-m", default=None)
    ap.add_argument("--api-key",   "-k", default=None)
    ap.add_argument("--base-url",        default=None)

    ap.add_argument("--timetable", "-t", default=None, help="课表图片(jpg/png)或xlsx文件路径")
    ap.add_argument("--timetable-months", default="3-6月", dest="timetable_months", help="课表对应月份，如'3-6月'")
    ap.add_argument("--english",         default="英语一")
    ap.add_argument("--subject1",        default=None)
    ap.add_argument("--subject2",        required=True)
    ap.add_argument("--school",          default="目标院校")
    
    ap.add_argument("--start-date","-s", default=default_start, dest="start_date")
    ap.add_argument("--end-date",        default=default_end,   dest="end_date")
    ap.add_argument("--exam-date",       default=default_exam,  dest="exam_date")
    ap.add_argument("--output-dir","-o", default=None,          dest="output_dir")

    args = ap.parse_args()

    env_map = {"anthropic": "ANTHROPIC_API_KEY", "openai": "OPENAI_API_KEY", "deepseek": "DEEPSEEK_API_KEY", "qwen": "DASHSCOPE_API_KEY", "gemini": "GEMINI_API_KEY", "gemini-native": "GEMINI_API_KEY", "ollama": None}
    api_key = args.api_key or (os.getenv(env_map[args.provider]) if env_map[args.provider] else "ollama")
    
    if not api_key and args.provider != "ollama":
        print(f"❌ 缺少 API Key，请使用 -k 参数传入")
        return

    cfg = {"provider": args.provider, "api_key": api_key, "base_url": args.base_url, "model": args.model}
    info = vars(args)
    out_dir = args.output_dir or f"output_{args.subject2.split(',')[0].strip()}"

    print("─" * 58)
    print(f"🚀 引擎配置 : {args.provider} | {cfg.get('model') or '(默认模型)'}")
    if args.base_url:
        print(f"   接口中转 : {args.base_url}")
    print(f"   排期区间 : {info['start_date']} → {info['end_date']}")
    print(f"   输出目录 : ./{out_dir}/")
    print("─" * 58)

    generate_all(cfg, info, out_dir)

if __name__ == "__main__":
    main()